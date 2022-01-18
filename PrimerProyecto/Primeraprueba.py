import time
import unittest
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

class registro_unittest(unittest.TestCase):
###MANEJADOR
    def setUp(self):
        self.driver = webdriver.Chrome(executable_path=r"C:\Users\user\Documents\Proyectos\drivernavegadores\chromedriver.exe")

##### ACCEDER AL REGISTRO DE USAURIO
    def test_registro_usuario(self):
        web = "https://alfogolarexpress.com/"
        driver = self.driver
        driver.get(web)
        driver.maximize_window()
        LinkRegister = driver.find_element_by_xpath("/html/body/header/div[1]/div/div/ul/li[1]/div/a[1]")
        LinkRegister.click()
        time.sleep(3)

### LLENADO DE FORMULARIO DE REGISTRO
### Hago uan variable para no escribir la funcion slf siempre que vaya a llamar el comando driver
        driver = self.driver
### Declaro una variable con la ruta del archivo Excel
        filesheet = "C:/Users/user/Documents/Proyectos/Selenium/Alfogolar/recursos/DatosLogin.xlsx"
#### Declaro una variable para leer el archivo excel con la funcion load_workbook
        wb = load_workbook(filesheet)
#### Valido que estoy visualizando la hoja correcta del archivo donde poseo los datos almacenados
        hojas = wb.get_sheet_names()
        print(hojas)
### Creo una variable para guardar los datos que estan en la hoja de excel
        datos = wb.get_sheet_by_name("Hoja1")

### Cierro el workbook

### Creo un ciclo for para ir chequeando cada fila y columna del archivo y a su vez guardarlos en las variables que estarab asignadas a los campos a llenar
        for i in range(1,21):
### Creo variables donde se guadaran cada uno de lso datos por columna y coloco el rango de letra que tenemos de datosw en el excel

            name, lastname, user, cod = datos[f'A{i}:D{i}'][0]
            ntlf, email, password = datos[f'E{i}:G{i}'][0]
            print(name.value, lastname.value, user.value, cod.value, ntlf.value, email.value, password.value)
            time.sleep (0.5)
        ### llenar formulario con datos sustraidos del excel

            driver.find_element_by_id("firstname").send_keys(name.value)
            time.sleep(0.25)
            driver.find_element_by_id("lastname").send_keys(lastname.value)
            time.sleep(0.25)
            driver.find_element_by_id("username").send_keys(user.value)
            time.sleep(0.25)

            select = driver.find_element_by_name("country_code")

### Seleccionar un datos de una lista select sin usar for
            #drop = Select(list)
            #drop.select_by_visible_text(cod.value)

### Seleccionar un dato de una lista select usando el ciclo for
            opcion = select.find_elements_by_tag_name("option")
            time.sleep(0.25)
            for i in opcion:
                print("Los valores son %s" % i.get_attribute("value"))
                i.click()
                time.sleep(0.00001)
            Seleccionar = Select(select)
            Seleccionar.select_by_value(str(cod.value))

            time.sleep(0.25)
            driver.find_element_by_name("mobile").send_keys(ntlf.value)
            driver.execute_script("window.scrollTo(0, 200);")
            driver.find_element_by_id("email").send_keys(email.value)
            time.sleep(0.25)
            driver.find_element_by_id("password").send_keys(password.value)
            time.sleep(0.25)
            driver.find_element_by_id("password-confirm").send_keys(password.value)
            time.sleep(0.25)

            driver.execute_script("window.scrollTo(0, 600);")
## Validar el capchat
            cap1 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[1]")
            cap1_text = cap1.text
            cap2 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[2]")
            cap2_text = cap2.text
            cap3 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[3]")
            cap3_text = cap3.text
            cap4 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[4]")
            cap4_text = cap4.text
            cap5 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[5]")
            cap5_text = cap5.text
            cap6 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/div/div/span[6]")
            cap6_text = cap6.text

            captchaname = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/div/form/div[9]/input")
            captchaname.send_keys(cap1_text + cap2_text + cap3_text + cap4_text + cap5_text + cap6_text)

### Aceptar terminos y condiciones

            driver.find_element_by_id("terms-conditions").click()
            time.sleep(0.25)
            driver.execute_script("window.scrollTo(0, 700);")
            aceptar = driver.find_element_by_id("terms-conditions").is_selected()
            for i in range(1, 3, +1):
                if aceptar == True:
                    driver.find_element_by_id("recaptcha").click()


                else :
                    print("Error debe aceptar terminos y condiciones")
                    driver.find_element_by_id("terms-conditions").click()

                break
            time.sleep(2)

            #visualizar si el registro fue exitoso y cerrar sesion
            login_exitoso = driver.find_elements_by_class_name("dashboard-menu-bar")
            if login_exitoso != "Iniciar Sesión":
                driver.find_element_by_xpath("/html/body/header/div[1]/div/div/ul/li[1]/div/a[2]").click()
                time.sleep(1)
                driver.find_element_by_xpath("/html/body/div[3]/div[7]/ul/li[7]/a").click()
            else:
                print("Error no se concretó el registro")


            time.sleep(1)
            LinkRegister = driver.find_element_by_xpath("/html/body/header/div[1]/div/div/ul/li[1]/div/a[1]")
            LinkRegister.click()


        print("Registros completados")
        wb.close()

    def tearDown(self):
        self.driver.close()


if __name__ == '__main__':
        unittest.main()
